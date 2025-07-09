from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from selenium import webdriver
from tkinter import messagebox
import tkinter as tk
import pandas as pd
import pyperclip
import time
import sys
import os


def iniciar_script():
    
    global diretorio 
    global nome_arquivo

    diretorio = entry_diretorio.get()
    nome_arquivo = entry_arquivo.get()

    if not diretorio:
        messagebox.showerror("Erro", "Cole o diret√≥rio do border√¥.")
        return

    if not nome_arquivo:
        messagebox.showerror("Erro", "Digite o nome do arquivo.")
        return

    root.destroy()
    executar_script(diretorio, nome_arquivo)


def executar_script(diretorio, nome_arquivo):

    # Configurar op√ß√µes do Chrome
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    # options.add_argument("--headless=new") # Executar em modo headless
    prefs = {
        "credentials_enable_service": False,  # Desativa o servi√ßo de credenciais
        "profile.password_manager_enabled": False  # Desativa o gerenciador de senhas
    }
    options.add_experimental_option("prefs", prefs)

    # Inicializar navegador com as op√ß√µes
    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico, options=options)

    def esperar_e_clicar_simples(driver, elemento_id, tempo_espera=30):
        """
        Espera at√© que o elemento identificado pelo ID seja clic√°vel e, ent√£o, realiza um duplo clique.

        :param driver: Inst√¢ncia do WebDriver.
        :param elemento_id: ID do elemento a ser clicado.
        :param tempo_espera: Tempo m√°ximo de espera em segundos (padr√£o: 30 segundos).
        """
        try:
            # Espera at√© que o elemento seja clic√°vel
            WebDriverWait(driver, tempo_espera).until(
                EC.element_to_be_clickable((By.ID, elemento_id))
            )

            # Localiza o elemento pelo ID
            elemento = driver.find_element(By.ID, elemento_id)

            # Realiza o duplo clique usando ActionChains
            action = ActionChains(driver)
            action.double_click(elemento).perform()  # Duplo clique
            print('Duplo clique realizado com sucesso.')
        except Exception as e:
            print(f"Erro ao realizar o duplo clique: {e}")

    def esperar_e_clicar(driver, elemento_id, tempo_espera=30, cliques=2):
        """
        Espera at√© que o elemento identificado pelo ID esteja presente e realiza m√∫ltiplos cliques.

        :param driver: Inst√¢ncia do WebDriver.
        :param elemento_id: ID do elemento a ser clicado.
        :param tempo_espera: Tempo m√°ximo de espera em segundos (padr√£o: 30 segundos).
        :param cliques: N√∫mero de cliques a serem realizados (padr√£o: 2 - duplo clique).
        """
        try:
            # Espera at√© que o elemento esteja presente no DOM
            elemento = WebDriverWait(driver, tempo_espera).until(
                EC.presence_of_element_located((By.ID, elemento_id))
            )

            # Espera at√© que o elemento esteja vis√≠vel e clic√°vel
            WebDriverWait(driver, tempo_espera).until(
                EC.element_to_be_clickable((By.ID, elemento_id))
            )

            # Executa os cliques necess√°rios
            action = ActionChains(driver)
            for _ in range(cliques):
                action.click(elemento)
            action.perform()
            
            print(f'{cliques} clique(s) realizado(s) com sucesso.')
        except Exception as e:
            print(f"Erro ao clicar no elemento '{elemento_id}': {e}")

    def inserir_Sem_Espa√ßo(driver, elemento_id, texto, tempo_espera=30):
        """
        Espera at√© que o elemento identificado pelo ID seja clic√°vel, limpa o campo e insere o texto.
        Se falhar, insere o texto for√ßadamente via JavaScript.
        Garante que n√£o haver√° espa√ßos extras ap√≥s a inser√ß√£o.

        :param driver: Inst√¢ncia do WebDriver.
        :param elemento_id: ID do elemento onde o texto ser√° inserido.
        :param texto: Texto a ser inserido no elemento.
        :param tempo_espera: Tempo m√°ximo de espera em segundos (padr√£o: 30 segundos).
        """
        texto = texto.strip()  # Remover qualquer espa√ßo em branco no in√≠cio e no final do texto
        
        try:
            # Espera at√© o elemento ficar clic√°vel
            elemento = WebDriverWait(driver, tempo_espera).until(
                EC.element_to_be_clickable((By.ID, elemento_id))
            )
            
            # Limpar o campo completamente com JavaScript
            driver.execute_script(f"document.getElementById('{elemento_id}').value = '';")
            print(f"Campo {elemento_id} limpo com sucesso.")

            # Agora, insira o texto sem espa√ßos extras
            driver.execute_script(f"document.getElementById('{elemento_id}').value = '{texto}';")
            
            # Garantir que o foco est√° no campo para a inser√ß√£o funcionar corretamente
            driver.execute_script(f"document.getElementById('{elemento_id}').focus();")
            
            print(f"Texto '{texto}' inserido com sucesso via JavaScript no elemento {elemento_id}.")
        except Exception as e:
            print(f"Erro ao inserir texto no elemento {elemento_id}: {e}")

            print(f"Erro ao inserir texto no elemento {elemento_id}: {e}")

    #Vari√°veis / Diret√≥rios / Loads
    caminho_arquivo = os.path.join(diretorio, nome_arquivo)
    wb = load_workbook(caminho_arquivo)
    ws = wb["Reembolso"]

    # Abrir p√°gina
    navegador.get("http://an148124.protheus.cloudtotvs.com.br:1703/webapp/")
    WebDriverWait(navegador, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    actions = ActionChains(navegador)

    #Logar e entrar no m√≥dulo manualmente
    time.sleep(60)

    #Outras a√ß√µes
    esperar_e_clicar_simples(navegador, "COMP4606")
    time.sleep(2)

    #Bordero
    esperar_e_clicar_simples(navegador, "COMP4614")
    time.sleep(2)

    #Bordero 2
    esperar_e_clicar_simples(navegador, "COMP4615")
    time.sleep(4)


    # ------ MONTAGEM DO BORDER√î -------------------#

    # Faz a leitura dos IDs de reembolsos na PLANILHA
    df = pd.read_excel(caminho_arquivo, sheet_name="Reembolso", engine="openpyxl")

    valores_coluna_f = df.iloc[:, 5] #F2 em diante

    valores_nao_vazios = valores_coluna_f.dropna().astype(int).apply(lambda x: f"{x:09d}")

    ids_planilha = valores_nao_vazios.tolist()

    fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid") #Grifar IDs que est√£o no border√¥

    ids_processados = set() #IDs j√° checados.

    while True:

        #Capturar n√∫mero do bordero
        elemento = WebDriverWait(navegador, 40).until(EC.presence_of_element_located((By.ID, "COMP6018")))
        actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
        actions.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()
        num_bordero = pyperclip.paste()

        # time.sleep(15) #Op√ß√£o para mudar a data do border√¥  

        #Banco
        inserir_Sem_Espa√ßo(navegador,"COMP6022", "756",30)
        time.sleep(1)

        #Agencia
        inserir_Sem_Espa√ßo(navegador,"COMP6023" ,"3337",30)
        time.sleep(1)

        #Conta
        inserir_Sem_Espa√ßo(navegador,"COMP6024", "3780624",30)
        time.sleep(1)

        #Modelo
        inserir_Sem_Espa√ßo(navegador,"COMP6030", "02",30)
        time.sleep(2)

        #Tipo Pagamento
        inserir_Sem_Espa√ßo(navegador,"COMP6031" ,"20",30)
        time.sleep(1)

        #OK
        esperar_e_clicar_simples(navegador, "COMP6032")
        time.sleep(7)

        #Scroll down para localizar o filtro
        tcbrowse = navegador.find_element(By.ID, "COMP6003")
        for _ in range(10):  
            tcbrowse.send_keys(Keys.PAGE_DOWN)
            time.sleep(0.5)

        #Selecinar o filtro
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(2)

        #Clicar no Confirmar
        esperar_e_clicar(navegador,"COMP6008")
        time.sleep(40)


        #Desmarcar todos os campos selecionados
        wa_element = navegador.find_element(By.ID, "COMP6008")
        navegador.execute_script("""
            const shadow = arguments[0].shadowRoot;
            const target = shadow.querySelector('th[id="0"]');
            if (target) {
                target.click();
            } else {
                console.error("Elemento <th id='0'> n√£o encontrado.");
            }
        """, wa_element)

        #Acessar o campo "Reembolso"
        time.sleep(10)
        actions.send_keys(Keys.ARROW_RIGHT).perform()
        time.sleep(4)
        actions.send_keys(Keys.ARROW_RIGHT).perform()
        time.sleep(4)
        actions.send_keys(Keys.ARROW_RIGHT).perform()
        time.sleep(4)

        #Vari√°veis de controle
        bordero = 0
        encontrados_nesse_bordero = 0
        tentativas_mesmo_id = 0
        ultimo_id_lido = None

        while bordero <= 99: #Tamanho do border√¥
            actions = ActionChains(navegador)
            actions.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()

            time.sleep(0.2)  # Pequeno delay para garantir que o clipboard atualize
            copied_text = pyperclip.paste()

            #Verifica se est√° repetindo o mesmo ID
            if copied_text == ultimo_id_lido:
                tentativas_mesmo_id +=1
            else:
                tentativas_mesmo_id = 0 #Reinicia o contador
                ultimo_id_lido = copied_text


            #Se repetiu o mesmo ID 3 vezes, salva o border√¥ e finaliza o script
            if tentativas_mesmo_id >= 3:

                #Salvar
                esperar_e_clicar_simples(navegador, "COMP6013")

                # Salvar altera√ß√µes Excel
                wb.save(caminho_arquivo)
                print(f"‚úÖ Border√¥ gerado com {encontrados_nesse_bordero} IDs.")
                time.sleep(3)
                

                print("\nüö´ Nenhum novo ID foi encontrado. Encerrando gera√ß√£o de border√¥s...")
                time.sleep(8)
                sys.exit()
                
            print("Reembolso Protheus:", copied_text)

            if copied_text in ids_planilha:
                ActionChains(navegador).send_keys(Keys.ENTER).perform()
                bordero +=1
                encontrados_nesse_bordero +=1
                ids_processados.add(copied_text)

                # Localiza o √≠ndice do ID copiado na lista e grifa de amarelo a c√©lula no Excel
                index = ids_planilha.index(copied_text) 
                excel_row = index + 2  # Linha real no Excel (F2 em diante)
                ws[f"F{excel_row}"].fill = fill_amarelo
                ws[f"I{excel_row}"] = num_bordero
                print(f"Colorindo F{excel_row} de amarelo.")

                
                print("ID Encontrado. T√≠tulos dentro do border√¥ no momento: ", {bordero})
                time.sleep(3)
                ActionChains(navegador).send_keys(Keys.ARROW_DOWN).perform()
                time.sleep(3)
            else:
                ActionChains(navegador).send_keys(Keys.ARROW_DOWN).perform()
                print("ID N√ÉO ENCONTRADO. Prosseguindo...")
                time.sleep(3)

        #Salvar
        esperar_e_clicar_simples(navegador, "COMP6013")
        time.sleep(3)

        # Salvar altera√ß√µes Excel
        wb.save(caminho_arquivo)
        time.sleep(3)

        # Verifica√ß√£o de fim de processamento
        if encontrados_nesse_bordero == 0:
            print("\nüö´ Nenhum novo ID foi encontrado. Encerrando gera√ß√£o de border√¥s...")
            time.sleep(7)
            sys.exit()
        else:
            print(f"‚úÖ Border√¥ gerado com {encontrados_nesse_bordero} IDs.")



root = tk.Tk()
root.title("Montagem de border√¥s")
root.geometry("300x150")
root.configure(bg="gray")

label_diretorio = tk.Label(root, text="Cole aqui o diret√≥rio do border√¥:", bg="gray", fg="white")
label_diretorio.pack(pady=3)
entry_diretorio = tk.Entry(root)
entry_diretorio.pack(pady=3)

label_arquivo = tk.Label(root, text="Digite o nome do arquivo:", bg="gray", fg="white")
label_arquivo.pack(pady=3)
entry_arquivo = tk.Entry(root)
entry_arquivo.pack(pady=3)

botao = tk.Button(root, text="Iniciar", command=iniciar_script, bg="darkgray", fg="white")
botao.pack(pady=10)

root.mainloop()
