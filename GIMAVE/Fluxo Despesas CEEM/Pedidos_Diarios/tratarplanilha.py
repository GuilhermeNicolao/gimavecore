import pandas as pd
from openpyxl import load_workbook

# Caminhos dos arquivos
caminho_pedido = r"C:\Users\Guilherme.Silva\Desktop\Pedido_Diario.xlsx"
caminho_grupo = r"C:\Users\Guilherme.Silva\Desktop\grupo.xlsx"

# Nome das abas
aba_pedido = "PedidoDiario"
aba_grupo = "GRUPO"

# Ler a planilha com pandas
df = pd.read_excel(caminho_pedido, sheet_name=aba_pedido)

# Carregar o arquivo Excel para edição
wb = load_workbook(caminho_pedido)
ws = wb[aba_pedido]

# Escrever "Grupo" na célula R1
ws.cell(row=1, column=18, value="Grupo")  # R1 = coluna 18

# Determinar o número de linhas para aplicar XLOOKUP
num_linhas = ws.max_row

# Criar a fórmula XLOOKUP referenciando 'grupo.xlsx'
caminho_grupo = caminho_grupo.replace("\\", "/")  # Excel usa '/' em caminhos
formula_base = f"=XLOOKUP(B{{}}, '[{caminho_grupo}]GRUPO'!$A:$A, '[{caminho_grupo}]GRUPO'!$B:$B, 0)"

# Inserir a fórmula na coluna R (a partir da linha 2)
for row in range(2, num_linhas + 1):
    formula = formula_base.format(row)
    ws[f'R{row}'].value = formula

# Salvar e fechar o arquivo
wb.save(caminho_pedido)
wb.close()

print("Edição concluída: Título adicionado e XLOOKUP inserido!")
